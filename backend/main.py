import os
import sys
import time
import random
import io
import re
import uuid
import asyncio
from datetime import datetime, timedelta
from typing import List, Optional

# Ensure backend directory is in python search path
current_dir = os.path.dirname(os.path.abspath(__file__))
if current_dir not in sys.path:
    sys.path.insert(0, current_dir)

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

import uvicorn
import openpyxl
from openpyxl import Workbook
from fastapi import FastAPI, Depends, HTTPException, Query, status, UploadFile, File, Body, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import desc, text

from database import engine, Base, get_db, SessionLocal
import models
import schemas

# ============================================================================
# FASTAPI APPLICATION SETUP
# ============================================================================
app = FastAPI(
    title="Coupon Management System API",
    description="Production-ready FastAPI backend with MySQL database for airline/travel coupon lifecycle management.",
    version="1.0.0"
)

# Enable CORS for frontend integration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ============================================================================
# DATABASE INITIALIZATION & SAMPLE DATA SEEDING
# ============================================================================
def init_database():
    """Create tables, ensure missing columns exist, and seed sample data if database is fresh"""
    try:
        Base.metadata.create_all(bind=engine)
        
        # Auto-ensure new columns exist if table was created previously without them
        try:
            with engine.connect() as conn:
                for table, col_name, col_type in [
                    ("bookings", "airline_pnr", "VARCHAR(50) DEFAULT NULL"),
                    ("bookings", "booking_type", "VARCHAR(50) DEFAULT NULL"),
                    ("bookings", "sector", "VARCHAR(150) DEFAULT NULL"),
                    ("bookings", "parent_pnr", "VARCHAR(50) DEFAULT NULL"),
                    ("bookings", "pax_name", "VARCHAR(150) DEFAULT NULL"),
                    ("bookings", "source_status", "VARCHAR(50) DEFAULT NULL"),
                    ("bookings", "username", "VARCHAR(100) DEFAULT NULL"),
                    ("coupon_rules", "office_id", "VARCHAR(50) DEFAULT NULL"),
                    ("coupon_rules", "booking_type", "VARCHAR(50) DEFAULT NULL"),
                ]:
                    try:
                        conn.execute(text(f"ALTER TABLE `{table}` ADD COLUMN `{col_name}` {col_type};"))
                        conn.commit()
                    except Exception:
                        pass
                
                # Normalize legacy Ticketed status to Confirmed
                try:
                    conn.execute(text("UPDATE `bookings` SET `status` = 'Confirmed' WHERE `status` = 'Ticketed';"))
                    conn.execute(text("UPDATE `bookings` SET `source_status` = 'Confirmed' WHERE `source_status` = 'Ticketed';"))
                    conn.commit()
                except Exception:
                    pass
        except Exception as e:
            print(f"[INFO] Column verification completed: {e}")

        db = SessionLocal()
        
        # Check if sample customers exist
        if db.query(models.Customer).count() == 0:
            print("[INFO] Seeding initial sample data into MySQL database...")
            
            # 1. Customers
            customers = [
                models.Customer(customer_id="CUST001", name="Rajesh Kumar", email="rajesh@email.com", phone="+91 9876543210", status="Active"),
                models.Customer(customer_id="CUST002", name="Priya Sharma", email="priya@email.com", phone="+91 9876543211", status="Active"),
                models.Customer(customer_id="CUST003", name="Amit Patel", email="amit@email.com", phone="+91 9876543212", status="Active"),
            ]
            db.add_all(customers)
            db.commit()

            # 2. Balances
            balances = [
                models.CouponBalance(customer_id="CUST001", total_earned=5000.0, pending=500.0, available=4500.0, redeemed=1000.0, expired=0.0, cancelled=0.0),
                models.CouponBalance(customer_id="CUST002", total_earned=3000.0, pending=0.0, available=3000.0, redeemed=0.0, expired=0.0, cancelled=0.0),
                models.CouponBalance(customer_id="CUST003", total_earned=2000.0, pending=200.0, available=1800.0, redeemed=0.0, expired=0.0, cancelled=0.0),
            ]
            db.add_all(balances)
            db.commit()

            # 3. Rules (Priority matching with Office ID & Booking Type)
            rules = [
                models.CouponRule(rule_id="RULE-001", office_id="OFF-DEL", booking_type="Round Trip", supplier="Supplier A", airline="IndiGo", fare_type="Super 6E", coupon_percent=3.0, priority=7, status="Active"),
                models.CouponRule(rule_id="RULE-002", office_id=None, booking_type="Round Trip", supplier="Supplier A", airline="IndiGo", fare_type=None, coupon_percent=2.0, priority=6, status="Active"),
                models.CouponRule(rule_id="RULE-003", office_id=None, booking_type=None, supplier="Supplier A", airline="Air India", fare_type="Flexi", coupon_percent=2.5, priority=5, status="Active"),
                models.CouponRule(rule_id="RULE-004", office_id=None, booking_type="One Way", supplier="Supplier B", airline="SpiceJet", fare_type=None, coupon_percent=1.5, priority=4, status="Active"),
                models.CouponRule(rule_id="RULE-005", office_id=None, booking_type=None, supplier=None, airline=None, fare_type=None, coupon_percent=1.0, priority=0, status="Active"),
            ]
            db.add_all(rules)
            db.commit()

            # 4. Bookings
            bookings = [
                models.Booking(booking_ref="BK-2024-001", customer_id="CUST001", supplier="Supplier A", airline="IndiGo", fare_type="Super 6E", booking_type="Round Trip", booking_fare=10000.0, booking_date=datetime(2024, 11, 20, 10, 0), travel_date=datetime(2024, 12, 10, 14, 30), status="Confirmed", source_status="Confirmed"),
                models.Booking(booking_ref="BK-2024-002", customer_id="CUST001", supplier="Supplier A", airline="Air India", fare_type="Flexi", booking_type="One Way", booking_fare=12000.0, booking_date=datetime(2024, 11, 22, 11, 30), travel_date=datetime(2024, 12, 20, 18, 0), status="Confirmed", source_status="Confirmed"),
                models.Booking(booking_ref="BK-2024-003", customer_id="CUST002", supplier="Supplier B", airline="SpiceJet", fare_type="Regular", booking_type="One Way", booking_fare=7500.0, booking_date=datetime(2024, 11, 25, 9, 15), travel_date=datetime(2024, 12, 15, 12, 0), status="Confirmed", source_status="Confirmed"),
            ]
            db.add_all(bookings)
            db.commit()

            # 5. Coupons
            coupons = [
                models.Coupon(coupon_id="CPN-BK-2024-001", booking_ref="BK-2024-001", customer_id="CUST001", coupon_percent=3.0, coupon_amount=300.0, status="Eligible", eligibility_date=datetime(2024, 12, 11, 14, 30)),
                models.Coupon(coupon_id="CPN-BK-2024-002", booking_ref="BK-2024-002", customer_id="CUST001", coupon_percent=2.5, coupon_amount=300.0, status="Eligible", eligibility_date=datetime(2024, 12, 21, 18, 0)),
                models.Coupon(coupon_id="CPN-BK-2024-003", booking_ref="BK-2024-003", customer_id="CUST002", coupon_percent=1.5, coupon_amount=112.5, status="Eligible", eligibility_date=datetime(2024, 12, 16, 12, 0)),
            ]
            db.add_all(coupons)
            db.commit()

            # 6. Redemptions
            redemptions = [
                models.CouponRedemption(redemption_id="RED-001", txn_id="RDM-1732796600.1", customer_id="CUST001", booking_ref="BK-2024-002", amount_redeemed=1000.0, booking_fare=12000.0, customer_payable=11000.0, status="Success", created_at=datetime(2024, 11, 28, 10, 0)),
                models.CouponRedemption(redemption_id="RED-002", txn_id="RDM-1732796600.2", customer_id="CUST002", booking_ref="BK-2024-003", amount_redeemed=500.0, booking_fare=7500.0, customer_payable=7000.0, status="Success", created_at=datetime(2024, 11, 28, 11, 30)),
            ]
            db.add_all(redemptions)
            db.commit()

            # 7. Ledger
            ledger = [
                models.CouponLedger(txn_id="TXN-1732796400.1", customer_id="CUST001", booking_ref="BK-2024-001", txn_type="Coupon Earned", booking_fare=10000.0, coupon_percent=3.0, coupon_earned=300.0, amount=300.0, status="Eligible", travel_date=datetime(2024, 12, 10, 14, 30), created_at=datetime(2024, 11, 20, 10, 0)),
                models.CouponLedger(txn_id="TXN-1732796400.2", customer_id="CUST001", booking_ref="BK-2024-002", txn_type="Coupon Earned", booking_fare=12000.0, coupon_percent=2.5, coupon_earned=300.0, amount=300.0, status="Eligible", travel_date=datetime(2024, 12, 20, 18, 0), created_at=datetime(2024, 11, 22, 11, 30)),
                models.CouponLedger(txn_id="RDM-1732796600.1", customer_id="CUST001", booking_ref="BK-2024-002", txn_type="Coupon Redeemed", booking_fare=12000.0, coupon_percent=0.0, coupon_earned=0.0, amount=-1000.0, status="Success", travel_date=None, created_at=datetime(2024, 11, 28, 10, 0)),
                models.CouponLedger(txn_id="TXN-1732796400.3", customer_id="CUST001", booking_ref="BK-2024-003", txn_type="Coupon Earned", booking_fare=15000.0, coupon_percent=3.0, coupon_earned=450.0, amount=450.0, status="Pending", travel_date=datetime(2024, 12, 28, 10, 0), created_at=datetime(2024, 11, 28, 11, 0)),
            ]
            db.add_all(ledger)
            db.commit()

            # 8. Settings
            settings = [
                models.SystemSetting(config_key="min_redemption", config_value="100"),
                models.SystemSetting(config_key="max_redemption", config_value="50000"),
                models.SystemSetting(config_key="expiry_days", config_value="365"),
                models.SystemSetting(config_key="allow_partial_redemption", config_value="true"),
                models.SystemSetting(config_key="allow_combined_offers", config_value="false"),
            ]
            db.add_all(settings)
            db.commit()
            print("[OK] Initial data seeded successfully!")

        db.close()
    except Exception as e:
        print(f"[WARNING] Database initialization error: {e}")

init_database()


# Helper function to generate unique transaction IDs
def gen_txn_id(prefix: str = "TXN") -> str:
    timestamp = int(time.time())
    suffix = uuid.uuid4().hex[:6].upper()
    return f"{prefix}-{timestamp}-{suffix}"


def compute_rule_priority(office_id=None, booking_type=None, supplier=None, airline=None, fare_type=None, explicit_priority=None) -> int:
    """
    Computes priority based on rule specificity:
    - Specific Office ID: +100
    - Specific Booking Type: +50
    - Specific Airline: +30
    - Specific Supplier: +20
    - Specific Fare Type: +10
    - Any Airline / Default Tier: 0
    """
    if explicit_priority is not None and int(explicit_priority) > 0:
        return int(explicit_priority)
    
    score = 0
    if office_id and str(office_id).strip() and str(office_id).strip().lower() not in ["any", "any office", "-"]:
        score += 100
    if booking_type and str(booking_type).strip() and str(booking_type).strip().lower() not in ["any", "any booking type", "any type", "any / standard", "-"]:
        score += 50
    if airline and str(airline).strip() and str(airline).strip().lower() not in ["any", "any airline", "all", "all airlines", "any / all", "-"]:
        score += 30
    if supplier and str(supplier).strip() and str(supplier).strip().lower() not in ["any", "any supplier", "-"]:
        score += 20
    if fare_type and str(fare_type).strip() and str(fare_type).strip().lower() not in ["any", "any fare", "any fare type", "any / standard", "-"]:
        score += 10
        
    return score


def find_matching_rule(rules, office_id=None, booking_type=None, supplier=None, airline=None, fare_type=None):
    """
    Finds the highest priority active rule that matches the booking criteria.
    Supports specific airlines (IndiGo, SpiceJet, Air India, etc.), specific fare types,
    and seamlessly falls back to 'Any Airline' / default tier if no specific rule matches.
    """
    req_off = (office_id or "").strip().lower()
    req_bt = (booking_type or "").strip().lower()
    req_s = (supplier or "").strip().lower()
    req_a = (airline or "").strip().lower()
    req_f = (fare_type or "").strip().lower()

    airline_aliases = {
        "6e": "indigo", "indigo": "indigo", "indigo airlines": "indigo",
        "ai": "air india", "airindia": "air india", "air india": "air india",
        "sg": "spicejet", "spicejet": "spicejet", "spice jet": "spicejet",
        "uk": "vistara", "vistara": "vistara",
        "qp": "akasa air", "akasa": "akasa air", "akasa air": "akasa air",
    }
    norm_req_a = airline_aliases.get(req_a, req_a)

    for rule in rules:
        if rule.status != "Active":
            continue
        r_off = (rule.office_id or "").strip().lower() if rule.office_id else ""
        r_bt = (rule.booking_type or "").strip().lower() if rule.booking_type else ""
        r_s = (rule.supplier or "").strip().lower() if rule.supplier else ""
        r_a = (rule.airline or "").strip().lower() if rule.airline else ""
        r_f = (rule.fare_type or "").strip().lower() if rule.fare_type else ""

        norm_r_a = airline_aliases.get(r_a, r_a)

        # Match Office ID
        match_off = (
            not r_off
            or not req_off
            or r_off in ["-", "any", "any office"]
            or req_off in ["-", "any", "any office"]
            or r_off == req_off
        )

        # Match Booking Type
        match_bt = (
            not r_bt
            or not req_bt
            or r_bt in ["-", "any", "any booking type", "any type", "any / standard"]
            or req_bt in ["-", "any", "any booking type", "any type", "any / standard"]
            or r_bt == req_bt
            or (req_bt in r_bt or r_bt in req_bt)
        )

        # Match Supplier
        match_s = (
            not r_s
            or not req_s
            or r_s in ["-", "any", "any supplier"]
            or req_s in ["-", "any", "any supplier"]
            or r_s == req_s
        )

        # Match Airline (IndiGo, SpiceJet, Air India, or Any Airline fallback)
        is_any_airline_rule = (not r_a or r_a in ["-", "any", "any airline", "all", "all airlines", "any / all"])
        is_any_airline_req = (not req_a or req_a in ["-", "any", "any airline", "all", "all airlines", "any / all"])
        if is_any_airline_rule:
            match_a = True
        elif is_any_airline_req:
            match_a = False
        else:
            match_a = (norm_r_a == norm_req_a or (norm_req_a and (norm_req_a in norm_r_a or norm_r_a in norm_req_a)))

        # Match Fare Type (Super 6E, Flexi, or Any Fare fallback)
        is_any_fare_rule = (not r_f or r_f in ["-", "any", "any fare", "any fare type", "any / standard"])
        is_any_fare_req = (not req_f or req_f in ["-", "any", "any fare", "any fare type", "any / standard"])
        if is_any_fare_rule:
            match_f = True
        elif is_any_fare_req:
            match_f = False
        else:
            match_f = (r_f == req_f or (req_f and (req_f in r_f or r_f in req_f)))

        if match_off and match_bt and match_s and match_a and match_f:
            return rule

    return None



# ============================================================================
# 1. HEALTH CHECK ENDPOINT
# ============================================================================
@app.get("/api/health", response_model=schemas.HealthResponse, tags=["Health"])
def health_check(db: Session = Depends(get_db)):
    """Check API server status and database connectivity"""
    try:
        db.execute(models.Customer.__table__.select().limit(1))
        return schemas.HealthResponse(
            status="healthy",
            timestamp=datetime.utcnow().isoformat(),
            database="connected",
            engine="MySQL 8.0"
        )
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={
                "status": "unhealthy",
                "timestamp": datetime.utcnow().isoformat(),
                "database": "error",
                "error": str(e)
            }
        )


# ============================================================================
# AUTOMATIC COUPON MATURITY HELPER (POST TRAVEL + 1 DAY)
# ============================================================================
def auto_mature_pending_coupons(db: Session, customer_id: Optional[str] = None) -> int:
    """
    High-performance batch maturity synchronization:
    - Performs a single indexed JOIN query instead of N+1 network queries.
    - If travel_date <= today: Transitions coupon to 'Eligible' and moves to 'Available'.
    - If travel_date > today: Reverts/maintains coupon in 'Pending'.
    """
    now = datetime.utcnow()
    today_date = now.date()

    query = db.query(models.Coupon, models.Booking).outerjoin(
        models.Booking, models.Coupon.booking_ref == models.Booking.booking_ref
    ).filter(models.Coupon.status.in_(["Pending", "Eligible", "Active"]))

    if customer_id:
        query = query.filter(models.Coupon.customer_id == customer_id)

    pairs = query.all()
    count = 0
    affected_customers = set()

    for coupon, booking in pairs:
        is_travel_completed = False
        if booking and booking.travel_date:
            b_tdate = booking.travel_date.date() if isinstance(booking.travel_date, datetime) else booking.travel_date
            if b_tdate <= today_date:
                is_travel_completed = True
        elif coupon.eligibility_date:
            c_edate = coupon.eligibility_date.date() if isinstance(coupon.eligibility_date, datetime) else coupon.eligibility_date
            if c_edate <= today_date:
                is_travel_completed = True

        if is_travel_completed:
            if coupon.status != "Eligible":
                coupon.status = "Eligible"
                coupon.updated_at = now
                affected_customers.add(coupon.customer_id)
                count += 1
        else:
            if coupon.status != "Pending":
                coupon.status = "Pending"
                coupon.updated_at = now
                affected_customers.add(coupon.customer_id)
                count += 1

    # Recalculate balances only for affected customers
    from sqlalchemy import func
    for cid in affected_customers:
        c_earned = db.query(func.sum(models.Coupon.coupon_amount)).filter(
            models.Coupon.customer_id == cid,
            models.Coupon.status.in_(["Pending", "Eligible", "Redeemed", "Active"])
        ).scalar() or 0.0

        c_pending = db.query(func.sum(models.Coupon.coupon_amount)).filter(
            models.Coupon.customer_id == cid,
            models.Coupon.status == "Pending"
        ).scalar() or 0.0

        c_eligible = db.query(func.sum(models.Coupon.coupon_amount)).filter(
            models.Coupon.customer_id == cid,
            models.Coupon.status == "Eligible"
        ).scalar() or 0.0

        c_redeemed = db.query(func.sum(models.CouponRedemption.amount_redeemed)).filter(
            models.CouponRedemption.customer_id == cid,
            models.CouponRedemption.status == "Success"
        ).scalar() or 0.0

        avail = max(0.0, float(c_eligible) - float(c_redeemed))

        bal = db.query(models.CouponBalance).filter(models.CouponBalance.customer_id == cid).first()
        if bal:
            bal.total_earned = float(c_earned)
            bal.pending = float(c_pending)
            bal.available = avail
            bal.redeemed = float(c_redeemed)
        else:
            db.add(models.CouponBalance(
                customer_id=cid,
                total_earned=float(c_earned),
                pending=float(c_pending),
                available=avail,
                redeemed=float(c_redeemed),
                expired=0.0,
                cancelled=0.0
            ))

    if count > 0:
        try:
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"[ERROR] auto_mature_pending_coupons commit failed: {e}")

    return count


def _run_sync_maturity_job():
    db = SessionLocal()
    try:
        matured = auto_mature_pending_coupons(db)
        if matured > 0:
            print(f"[AUTO-MATURITY] Auto-credited {matured} coupon(s) for completed travel.")
    except Exception as e:
        print(f"[AUTO-MATURITY-ERROR] {e}")
    finally:
        db.close()


async def background_auto_maturity_worker():
    """Non-blocking background worker running in a separate thread every 60 seconds"""
    while True:
        try:
            await asyncio.to_thread(_run_sync_maturity_job)
        except Exception as e:
            print(f"[WORKER-ERROR] {e}")
        await asyncio.sleep(60)


@app.on_event("startup")
async def start_background_auto_maturity():
    """Start background auto-maturity worker on server startup"""
    asyncio.create_task(background_auto_maturity_worker())


# ============================================================================
# DASHBOARD STATS ENDPOINT
# ============================================================================
@app.get("/api/dashboard/stats", tags=["Dashboard"])
def get_dashboard_stats(db: Session = Depends(get_db)):
    """Fetch system-wide aggregated metrics, recent transactions, and health for the dashboard"""
    try:
        from sqlalchemy import func
        total_customers = db.query(func.count(models.Customer.customer_id)).scalar() or 0
        total_bookings = db.query(func.count(models.Booking.booking_ref)).scalar() or 0
        total_rules = db.query(func.count(models.CouponRule.rule_id)).filter(models.CouponRule.status == "Active").scalar() or 0
        
        # Aggregate balances
        balance_agg = db.query(
            func.sum(models.CouponBalance.available).label("available"),
            func.sum(models.CouponBalance.pending).label("pending"),
            func.sum(models.CouponBalance.total_earned).label("total_earned"),
            func.sum(models.CouponBalance.redeemed).label("redeemed"),
            func.sum(models.CouponBalance.expired).label("expired"),
            func.sum(models.CouponBalance.cancelled).label("cancelled"),
        ).first()

        available = float(balance_agg.available or 0.0) if balance_agg else 0.0
        pending = float(balance_agg.pending or 0.0) if balance_agg else 0.0
        total_earned = float(balance_agg.total_earned or 0.0) if balance_agg else 0.0
        redeemed = float(balance_agg.redeemed or 0.0) if balance_agg else 0.0
        expired = float(balance_agg.expired or 0.0) if balance_agg else 0.0

        # Recent transactions (indexed by primary key id)
        recent = db.query(models.CouponLedger).order_by(desc(models.CouponLedger.id)).limit(8).all()
        recent_txns = [
            {
                "txn_id": t.txn_id,
                "customer_id": t.customer_id,
                "booking_ref": t.booking_ref,
                "type": t.txn_type,
                "amount": float(t.amount or 0.0),
                "status": t.status,
                "date": t.created_at.isoformat() if t.created_at else datetime.utcnow().isoformat()
            }
            for t in recent
        ]

        return {
            "status": "success",
            "available": available,
            "pending": pending,
            "total_earned": total_earned,
            "redeemed": redeemed,
            "expired": expired,
            "total_customers": total_customers,
            "total_bookings": total_bookings,
            "active_rules": total_rules,
            "recent_transactions": recent_txns,
            "engine": "MySQL 8.0"
        }
    except Exception as e:
        return {
            "status": "error",
            "available": 0.0,
            "pending": 0.0,
            "total_earned": 0.0,
            "redeemed": 0.0,
            "expired": 0.0,
            "total_customers": 0,
            "total_bookings": 0,
            "active_rules": 0,
            "recent_transactions": [],
            "error": str(e)
        }


# ============================================================================
# CLIENT PORTAL DASHBOARD ENDPOINT
# ============================================================================
@app.get("/api/client/dashboard/{customer_id}", tags=["Client Portal"])
def get_client_dashboard(customer_id: str, db: Session = Depends(get_db)):
    """Fetch personalized client dashboard: profile, balance, bookings, coupons, ledger"""
    customer_id = customer_id.strip()
    
    # Auto mature any pending coupons whose travel date + 1 day has passed for this customer
    auto_mature_pending_coupons(db, customer_id=customer_id)

    customer = db.query(models.Customer).filter(models.Customer.customer_id == customer_id).first()
    if not customer:
        customer = models.Customer(
            customer_id=customer_id,
            name=f"Client {customer_id}",
            email=f"{customer_id.lower()}@travel.com",
            status="Active"
        )
        db.add(customer)
        db.commit()
        db.refresh(customer)

    # Balance
    balance = db.query(models.CouponBalance).filter(models.CouponBalance.customer_id == customer_id).first()
    if not balance:
        balance = models.CouponBalance(
            customer_id=customer_id,
            total_earned=0.0,
            pending=0.0,
            available=0.0,
            redeemed=0.0,
            expired=0.0,
            cancelled=0.0
        )
        db.add(balance)
        db.commit()
        db.refresh(balance)

    # Bookings
    bookings = db.query(models.Booking).filter(
        models.Booking.customer_id == customer_id
    ).order_by(desc(models.Booking.booking_date)).limit(50).all()

    # Coupons
    coupons = db.query(models.Coupon).filter(
        models.Coupon.customer_id == customer_id
    ).order_by(desc(models.Coupon.created_at)).limit(50).all()

    # Ledger
    ledger = db.query(models.CouponLedger).filter(
        models.CouponLedger.customer_id == customer_id
    ).order_by(desc(models.CouponLedger.created_at)).limit(25).all()

    return {
        "status": "success",
        "client": {
            "customer_id": customer.customer_id,
            "name": customer.name,
            "email": customer.email,
            "phone": customer.phone or "-",
            "status": customer.status
        },
        "balance": {
            "available": float(balance.available or 0.0),
            "pending": float(balance.pending or 0.0),
            "total_earned": float(balance.total_earned or 0.0),
            "redeemed": float(balance.redeemed or 0.0),
            "expired": float(balance.expired or 0.0)
        },
        "bookings_count": len(bookings),
        "coupons_count": len(coupons),
        "recent_bookings": [
            {
                "booking_ref": b.booking_ref,
                "pax_name": b.pax_name or "-",
                "sector": b.sector or "-",
                "airline_pnr": b.airline_pnr or "-",
                "booking_type": b.booking_type or "-",
                "booking_fare": float(b.booking_fare or 0.0),
                "travel_date": b.travel_date.strftime("%Y-%m-%d") if b.travel_date else "-",
                "booked_date": b.booking_date.strftime("%Y-%m-%d %H:%M") if b.booking_date else "-",
                "status": b.source_status or b.status or "Confirmed"
            }
            for b in bookings
        ],
        "coupons": [
            {
                "coupon_id": c.coupon_id,
                "booking_ref": c.booking_ref,
                "amount": float(c.coupon_amount or 0.0),
                "percent": float(c.coupon_percent or 1.0),
                "status": c.status,
                "eligibility_date": c.eligibility_date.strftime("%Y-%m-%d") if c.eligibility_date else "-"
            }
            for c in coupons
        ],
        "ledger": [
            {
                "txn_id": l.txn_id,
                "booking_ref": l.booking_ref,
                "type": l.txn_type,
                "amount": float(l.amount or 0.0),
                "status": l.status,
                "date": l.created_at.strftime("%Y-%m-%d") if l.created_at else "-"
            }
            for l in ledger
        ]
    }


# ============================================================================
# 2. GET CUSTOMER BALANCE
# ============================================================================
@app.get("/api/coupon/balance/{customer_id}", response_model=schemas.BalanceResponse, tags=["Coupons"])
def get_customer_balance(customer_id: str, db: Session = Depends(get_db)):
    """Fetch real-time coupon balance breakdown for a specific customer"""
    customer_id = customer_id.strip()
    
    # Auto mature any pending coupons that passed travel_date + 1 day
    auto_mature_pending_coupons(db, customer_id=customer_id)

    balance = db.query(models.CouponBalance).filter(models.CouponBalance.customer_id == customer_id).first()
    
    if not balance:
        cust = db.query(models.Customer).filter(models.Customer.customer_id == customer_id).first()
        if not cust:
            cust = models.Customer(
                customer_id=customer_id,
                name=f"Customer {customer_id}",
                email=f"{customer_id.lower()}@example.com",
                status="Active"
            )
            db.add(cust)
            db.commit()
            
        balance = models.CouponBalance(
            customer_id=customer_id,
            total_earned=0.0,
            pending=0.0,
            available=0.0,
            redeemed=0.0,
            expired=0.0,
            cancelled=0.0
        )
        db.add(balance)
        db.commit()
        db.refresh(balance)

    return schemas.BalanceResponse(
        customer_id=balance.customer_id,
        total_earned=float(balance.total_earned or 0.0),
        pending=float(balance.pending or 0.0),
        available=float(balance.available or 0.0),
        redeemed=float(balance.redeemed or 0.0),
        expired=float(balance.expired or 0.0),
        cancelled=float(balance.cancelled or 0.0)
    )


# ============================================================================
# 3. EARN COUPON (POST /api/coupon/earn)
# ============================================================================
@app.post("/api/coupon/earn", response_model=schemas.EarnCouponResponse, tags=["Coupons"])
def earn_coupon(payload: schemas.EarnCouponRequest, db: Session = Depends(get_db)):
    """
    Calculate and create a pending coupon for a booking based on rule hierarchy.
    Supports matching on Office ID, Booking Type, Supplier, Airline, and Fare Type.
    """
    rules = db.query(models.CouponRule).filter(models.CouponRule.status == "Active").order_by(
        desc(models.CouponRule.priority),
        models.CouponRule.id.asc()
    ).all()

    matched_rule = find_matching_rule(
        rules=rules,
        office_id=payload.office_id,
        booking_type=payload.booking_type,
        supplier=payload.supplier,
        airline=payload.airline,
        fare_type=payload.fare_type
    )

    matched_percent = float(matched_rule.coupon_percent) if matched_rule else 1.0

    fare = float(payload.booking_fare or 0.0)
    percent = float(payload.percent if payload.percent is not None else (payload.coupon_percent if payload.coupon_percent is not None else matched_percent))
    max_cap = float(payload.fare_limit if payload.fare_limit is not None else (payload.flat_amount or 0.0))

    # 1. Percentage calculation
    calc_amount = round(fare * (percent / 100.0), 2)

    # 2. Apply Max Discount Cap (Fare Limit) if specified
    if max_cap > 0:
        capped_amount = min(calc_amount, max_cap)
    else:
        capped_amount = calc_amount

    # 3. Actual discount / coupon (cannot exceed ticket fare)
    coupon_earned = round(min(fare, capped_amount), 2)
    booking_fare = fare
    matched_percent = percent

    customer = db.query(models.Customer).filter(models.Customer.customer_id == payload.customer_id).first()
    if not customer:
        customer = models.Customer(
            customer_id=payload.customer_id,
            name=payload.customer_name or f"Customer {payload.customer_id}",
            email=f"{payload.customer_id.lower()}@example.com",
            status="Active"
        )
        db.add(customer)
        db.commit()

    try:
        travel_dt = datetime.fromisoformat(payload.travel_date.replace("Z", "+00:00")).replace(tzinfo=None)
    except Exception:
        travel_dt = datetime.utcnow() + timedelta(days=14)

    try:
        booking_dt = datetime.fromisoformat(payload.booking_date.replace("Z", "+00:00")).replace(tzinfo=None) if payload.booking_date else datetime.utcnow()
    except Exception:
        booking_dt = datetime.utcnow()

    eligibility_dt = travel_dt + timedelta(days=1)

    booking = db.query(models.Booking).filter(models.Booking.booking_ref == payload.booking_ref).first()
    if not booking:
        booking = models.Booking(
            booking_ref=payload.booking_ref,
            customer_id=payload.customer_id,
            supplier=payload.supplier,
            airline=payload.airline,
            fare_type=payload.fare_type,
            booking_type=payload.booking_type,
            booking_fare=booking_fare,
            booking_date=booking_dt,
            travel_date=travel_dt,
            status="Confirmed",
            source_status="Confirmed"
        )
        db.add(booking)
    else:
        booking.booking_fare = booking_fare
        booking.travel_date = travel_dt
        booking.booking_type = payload.booking_type or booking.booking_type
        booking.status = "Confirmed"

    coupon_id = f"CPN-{payload.booking_ref}"
    existing_coupon = db.query(models.Coupon).filter(models.Coupon.coupon_id == coupon_id).first()
    if existing_coupon:
        existing_coupon.coupon_amount = coupon_earned
        existing_coupon.coupon_percent = matched_percent
        existing_coupon.status = "Pending"
        existing_coupon.eligibility_date = eligibility_dt
    else:
        new_coupon = models.Coupon(
            coupon_id=coupon_id,
            booking_ref=payload.booking_ref,
            customer_id=payload.customer_id,
            coupon_percent=matched_percent,
            coupon_amount=coupon_earned,
            status="Pending",
            eligibility_date=eligibility_dt,
            expiry_date=eligibility_dt + timedelta(days=365)
        )
        db.add(new_coupon)

    balance = db.query(models.CouponBalance).filter(models.CouponBalance.customer_id == payload.customer_id).first()
    if not balance:
        balance = models.CouponBalance(
            customer_id=payload.customer_id,
            total_earned=coupon_earned,
            pending=coupon_earned,
            available=0.0,
            redeemed=0.0,
            expired=0.0,
            cancelled=0.0
        )
        db.add(balance)
    else:
        balance.total_earned = float(balance.total_earned or 0.0) + coupon_earned
        balance.pending = float(balance.pending or 0.0) + coupon_earned

    txn_id = gen_txn_id("TXN")
    ledger_entry = models.CouponLedger(
        txn_id=txn_id,
        customer_id=payload.customer_id,
        booking_ref=payload.booking_ref,
        txn_type="Coupon Earned",
        booking_fare=booking_fare,
        coupon_percent=matched_percent,
        coupon_earned=coupon_earned,
        amount=coupon_earned,
        status="Pending",
        travel_date=travel_dt,
        created_at=datetime.utcnow()
    )
    db.add(ledger_entry)

    db.commit()

    return schemas.EarnCouponResponse(
        status="success",
        txn_id=txn_id,
        coupon_earned=coupon_earned,
        coupon_percent=matched_percent,
        coupon_status="Pending",
        eligibility_date=eligibility_dt.isoformat(),
        message="Coupon will be available after travel completion"
    )


# ============================================================================
# 4. RELEASE COUPON (POST /api/coupon/release)
# ============================================================================
@app.post("/api/coupon/release", response_model=schemas.ReleaseCouponResponse, tags=["Coupons"])
def release_coupon(payload: schemas.ReleaseCouponRequest, db: Session = Depends(get_db)):
    """
    Move coupon from Pending to Eligible after travel completion.
    Only a coupon that is CURRENTLY Pending can be released — prevents duplicate/double release.
    """
    coupon = db.query(models.Coupon).filter(
        models.Coupon.booking_ref == payload.booking_ref
    ).first()

    if not coupon:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No coupon found for booking reference: {payload.booking_ref}"
        )

    if coupon.status != "Pending":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot release — coupon for booking '{payload.booking_ref}' is already '{coupon.status}', not Pending. "
                   f"(Prevents duplicate release / double-counting the balance.)"
        )

    customer_id = coupon.customer_id
    coupon_amount = float(coupon.coupon_amount)
    coupon.status = "Eligible"

    balance = db.query(models.CouponBalance).filter(models.CouponBalance.customer_id == customer_id).first()
    if balance:
        balance.pending = max(0.0, float(balance.pending or 0.0) - coupon_amount)
        balance.available = float(balance.available or 0.0) + coupon_amount
    else:
        balance = models.CouponBalance(
            customer_id=customer_id,
            total_earned=coupon_amount,
            pending=0.0,
            available=coupon_amount,
            redeemed=0.0,
            expired=0.0,
            cancelled=0.0
        )
        db.add(balance)

    txn_id = gen_txn_id("RLS")
    ledger_entry = models.CouponLedger(
        txn_id=txn_id,
        customer_id=customer_id,
        booking_ref=payload.booking_ref,
        txn_type="Coupon Released",
        booking_fare=0.0,
        coupon_percent=0.0,
        coupon_earned=0.0,
        amount=coupon_amount,
        status="Eligible",
        travel_date=datetime.utcnow(),
        created_at=datetime.utcnow()
    )
    db.add(ledger_entry)
    db.commit()

    return schemas.ReleaseCouponResponse(
        status="success",
        txn_id=txn_id,
        coupon_amount=coupon_amount,
        new_status="Eligible"
    )


# ============================================================================
# 4B. AUTO MATURE TRIGGER (POST /api/coupon/auto-mature)
# ============================================================================
@app.post("/api/coupon/auto-mature", tags=["Coupons"])
def trigger_auto_maturity(customer_id: Optional[str] = None, db: Session = Depends(get_db)):
    """
    Trigger automatic coupon maturation for all eligible pending coupons (post travel + 1 day).
    """
    count = auto_mature_pending_coupons(db, customer_id=customer_id)
    return {
        "status": "success",
        "matured_count": count,
        "message": f"Successfully processed {count} matured coupon(s) and credited to available balance."
    }


# ============================================================================
# 5. REDEEM COUPON (POST /api/coupon/redeem)
# ============================================================================
@app.post("/api/coupon/redeem", response_model=schemas.RedeemCouponResponse, tags=["Coupons"])
def redeem_coupon(payload: schemas.RedeemCouponRequest, db: Session = Depends(get_db)):
    """
    Redeem available coupon balance against a booking fare.
    """
    amount_to_redeem = float(payload.amount_to_redeem)
    booking_fare = float(payload.booking_fare)

    if amount_to_redeem <= 0:
        raise HTTPException(status_code=400, detail="Redemption amount must be greater than 0")

    balance = db.query(models.CouponBalance).filter(models.CouponBalance.customer_id == payload.customer_id).first()
    if not balance or float(balance.available or 0.0) < amount_to_redeem:
        available_amt = float(balance.available) if balance else 0.0
        raise HTTPException(
            status_code=400,
            detail=f"Insufficient available coupon balance. Available: ₹{available_amt:.2f}, Requested: ₹{amount_to_redeem:.2f}"
        )

    min_setting = db.query(models.SystemSetting).filter(models.SystemSetting.config_key == "min_redemption").first()
    if min_setting:
        try:
            min_val = float(min_setting.config_value)
            if amount_to_redeem < min_val and amount_to_redeem < float(balance.available):
                raise HTTPException(status_code=400, detail=f"Minimum redemption amount is {min_val:.2f}")
        except ValueError:
            pass

    customer_payable = max(0.0, round(booking_fare - amount_to_redeem, 2))

    balance.available = float(balance.available) - amount_to_redeem
    balance.redeemed = float(balance.redeemed or 0.0) + amount_to_redeem

    txn_id = gen_txn_id("RDM")
    redemption_id = f"RED-{payload.customer_id}-{payload.booking_ref}-{int(time.time()) % 100000}"

    redemption = models.CouponRedemption(
        redemption_id=redemption_id,
        txn_id=txn_id,
        customer_id=payload.customer_id,
        booking_ref=payload.booking_ref,
        amount_redeemed=amount_to_redeem,
        booking_fare=booking_fare,
        customer_payable=customer_payable,
        status="Success",
        created_at=datetime.utcnow()
    )
    db.add(redemption)

    ledger_entry = models.CouponLedger(
        txn_id=txn_id,
        customer_id=payload.customer_id,
        booking_ref=payload.booking_ref,
        txn_type="Coupon Redeemed",
        booking_fare=booking_fare,
        coupon_percent=0.0,
        coupon_earned=0.0,
        amount=-amount_to_redeem,
        status="Success",
        travel_date=None,
        created_at=datetime.utcnow()
    )
    db.add(ledger_entry)

    db.commit()
    db.refresh(balance)

    return schemas.RedeemCouponResponse(
        status="success",
        redemption_id=redemption_id,
        txn_id=txn_id,
        coupon_redeemed=amount_to_redeem,
        booking_fare=booking_fare,
        customer_payable=customer_payable,
        remaining_coupon_balance=float(balance.available)
    )


# ============================================================================
# 6. REVERSE COUPON (POST /api/coupon/reverse)
# ============================================================================
@app.post("/api/coupon/reverse", response_model=schemas.ReverseCouponResponse, tags=["Coupons"])
def reverse_coupon(payload: schemas.ReverseCouponRequest, db: Session = Depends(get_db)):
    """
    Cancel or reverse a coupon due to booking cancellation, refund, or void.
    """
    coupon = db.query(models.Coupon).filter(models.Coupon.booking_ref == payload.original_booking_ref).first()
    
    if not coupon:
        ledger = db.query(models.CouponLedger).filter(
            models.CouponLedger.booking_ref == payload.original_booking_ref
        ).order_by(desc(models.CouponLedger.id)).first()

        if not ledger:
            raise HTTPException(
                status_code=404,
                detail=f"No coupon or transaction found for booking ref: {payload.original_booking_ref}"
            )
        
        customer_id = ledger.customer_id
        coupon_amount = float(ledger.coupon_earned or abs(ledger.amount))
        current_status = ledger.status
    else:
        customer_id = coupon.customer_id
        coupon_amount = float(coupon.coupon_amount)
        current_status = coupon.status
        coupon.status = "Reversed"
        coupon.remarks = f"{payload.reason}: {payload.remarks or ''}".strip()

    balance = db.query(models.CouponBalance).filter(models.CouponBalance.customer_id == customer_id).first()
    if balance:
        if current_status == "Pending":
            balance.pending = max(0.0, float(balance.pending or 0.0) - coupon_amount)
            balance.total_earned = max(0.0, float(balance.total_earned or 0.0) - coupon_amount)
        elif current_status == "Eligible":
            balance.available = max(0.0, float(balance.available or 0.0) - coupon_amount)
            balance.total_earned = max(0.0, float(balance.total_earned or 0.0) - coupon_amount)
        
        balance.cancelled = float(balance.cancelled or 0.0) + coupon_amount

    txn_id = gen_txn_id("REV")
    ledger_entry = models.CouponLedger(
        txn_id=txn_id,
        customer_id=customer_id,
        booking_ref=payload.original_booking_ref,
        txn_type="Coupon Reversed",
        booking_fare=0.0,
        coupon_percent=0.0,
        coupon_earned=0.0,
        amount=-coupon_amount,
        status="Reversed",
        travel_date=None,
        created_at=datetime.utcnow()
    )
    db.add(ledger_entry)
    db.commit()

    return schemas.ReverseCouponResponse(
        status="success",
        txn_id=txn_id,
        action="Reversed",
        coupon_amount=coupon_amount,
        message=f"Coupon reversed successfully ({payload.reason})"
    )


# ============================================================================
# 7. GET TRANSACTION LEDGER (GET /api/coupon/ledger/{customer_id})
# ============================================================================
@app.get("/api/coupon/ledger", response_model=List[schemas.LedgerItem], tags=["Coupons"])
def get_all_coupon_ledger(db: Session = Depends(get_db)):
    """Fetch transaction ledger entries across ALL customers (for the Ledger dashboard view)."""
    entries = db.query(models.CouponLedger).order_by(desc(models.CouponLedger.created_at)).all()

    return [
        schemas.LedgerItem(
            txn_id=e.txn_id,
            customer_id=e.customer_id,
            booking_ref=e.booking_ref,
            type=e.txn_type,
            txn_type=e.txn_type,
            booking_fare=float(e.booking_fare or 0.0),
            coupon_percent=float(e.coupon_percent or 0.0),
            coupon_earned=float(e.coupon_earned or 0.0),
            amount=float(e.amount or 0.0),
            status=e.status,
            date=e.created_at.isoformat() if e.created_at else datetime.utcnow().isoformat(),
            travel_date=e.travel_date.isoformat() if e.travel_date else None
        )
        for e in entries
    ]


@app.get("/api/coupon/ledger/{customer_id}", response_model=schemas.LedgerResponse, tags=["Coupons"])
def get_coupon_ledger(customer_id: str, db: Session = Depends(get_db)):
    """Fetch complete transaction history and audit trail for a customer"""
    entries = db.query(models.CouponLedger).filter(
        models.CouponLedger.customer_id == customer_id
    ).order_by(desc(models.CouponLedger.created_at)).all()

    ledger_items = []
    for e in entries:
        ledger_items.append(
            schemas.LedgerItem(
                txn_id=e.txn_id,
                customer_id=e.customer_id,
                booking_ref=e.booking_ref,
                type=e.txn_type,
                txn_type=e.txn_type,
                booking_fare=float(e.booking_fare or 0.0),
                coupon_percent=float(e.coupon_percent or 0.0),
                coupon_earned=float(e.coupon_earned or 0.0),
                amount=float(e.amount or 0.0),
                status=e.status,
                date=e.created_at.isoformat() if e.created_at else datetime.utcnow().isoformat(),
                travel_date=e.travel_date.isoformat() if e.travel_date else None
            )
        )

    return schemas.LedgerResponse(
        customer_id=customer_id,
        ledger=ledger_items
    )


# ============================================================================
# 8. CRUD ENDPOINTS FOR UI INTEGRATION (Customers, Bookings, Rules, etc.)
# ============================================================================

# --- Customers ---
@app.get("/api/customers", response_model=List[schemas.CustomerResponse], tags=["Management"])
def list_customers(db: Session = Depends(get_db)):
    """Get all registered customers"""
    customers = db.query(models.Customer).all()
    return [
        schemas.CustomerResponse(
            id=c.customer_id,
            customer_id=c.customer_id,
            name=c.name,
            email=c.email,
            phone=c.phone,
            status=c.status,
            created_at=c.created_at.isoformat() if c.created_at else None
        )
        for c in customers
    ]

@app.post("/api/customers", response_model=schemas.CustomerResponse, tags=["Management"])
def create_customer(payload: schemas.CustomerCreate, db: Session = Depends(get_db)):
    """Create a new customer and initialize balance in MySQL"""
    cust_id = payload.id or payload.customer_id
    if not cust_id:
        cust_id = f"CUST{random.randint(100, 999)}"

    existing = db.query(models.Customer).filter(models.Customer.customer_id == cust_id).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Customer ID {cust_id} already exists")

    customer = models.Customer(
        customer_id=cust_id,
        name=payload.name,
        email=payload.email,
        phone=payload.phone,
        status=payload.status or "Active"
    )
    db.add(customer)
    
    balance = models.CouponBalance(
        customer_id=cust_id,
        total_earned=0.0,
        pending=0.0,
        available=0.0,
        redeemed=0.0,
        expired=0.0,
        cancelled=0.0
    )
    db.add(balance)
    db.commit()
    db.refresh(customer)

    return schemas.CustomerResponse(
        id=customer.customer_id,
        customer_id=customer.customer_id,
        name=customer.name,
        email=customer.email,
        phone=customer.phone,
        status=customer.status,
        created_at=customer.created_at.isoformat() if customer.created_at else None
    )


# --- Bookings ---
@app.get("/api/bookings", response_model=List[schemas.BookingResponse], tags=["Management"])
def list_bookings(db: Session = Depends(get_db)):
    """Get all bookings with full passenger, PNR, and travel details"""
    bookings = db.query(models.Booking).order_by(desc(models.Booking.booking_date)).all()
    res = []
    for b in bookings:
        client_name = b.customer.name if b.customer else f"Customer {b.customer_id}"
        res.append(schemas.BookingResponse(
            ref=b.booking_ref,
            booking_ref=b.booking_ref,
            customer=b.customer_id,
            customer_id=b.customer_id,
            customer_name=client_name,
            client=client_name,
            supplier=b.supplier,
            airline=b.airline,
            fare_type=b.fare_type,
            fare=float(b.booking_fare or 0.0),
            booking_fare=float(b.booking_fare or 0.0),
            net_amount=float(b.booking_fare or 0.0),
            travel=b.travel_date.strftime("%Y-%m-%d") if b.travel_date else "",
            travel_date=b.travel_date.isoformat() if b.travel_date else "",
            booking_date=b.booking_date.strftime("%Y-%m-%d %H:%M") if b.booking_date else "",
            booked_date=b.booking_date.strftime("%Y-%m-%d %H:%M") if b.booking_date else "",
            status=b.status,
            airline_pnr=b.airline_pnr,
            booking_type=b.booking_type,
            sector=b.sector,
            parent_pnr=b.parent_pnr,
            pax_name=b.pax_name,
            source_status=b.source_status,
            username=b.username
        ))
    return res

@app.post("/api/bookings", response_model=schemas.BookingResponse, tags=["Management"])
def create_booking(payload: schemas.BookingCreate, db: Session = Depends(get_db)):
    """Create a new booking with passenger and travel details"""
    ref = payload.ref or payload.booking_ref
    cust_id = payload.customer or payload.customer_id or payload.client
    cust_name = payload.customer_name or payload.client or f"Customer {cust_id}"
    fare = payload.fare if payload.fare is not None else (payload.booking_fare if payload.booking_fare is not None else (payload.net_amount or 0.0))

    if not ref or not cust_id:
        raise HTTPException(status_code=400, detail="Booking reference (S PNR) and Customer / Client ID are required")

    cust = db.query(models.Customer).filter(models.Customer.customer_id == cust_id).first()
    if not cust:
        cust = models.Customer(
            customer_id=cust_id,
            name=cust_name,
            email=f"{cust_id.lower().replace(' ', '_')}@example.com",
            status="Active"
        )
        db.add(cust)
        db.commit()
    elif cust_name and cust.name.startswith("Customer ") and cust_name != f"Customer {cust_id}":
        cust.name = cust_name
        db.commit()

    travel_raw = payload.travel_date or payload.travel
    travel_dt = parse_flexible_date(travel_raw, default=datetime.utcnow() + timedelta(days=7))

    booking_date_raw = payload.booking_date or payload.booked_date
    booking_dt = parse_flexible_date(booking_date_raw, default=datetime.utcnow())

    booking = models.Booking(
        booking_ref=ref,
        customer_id=cust_id,
        supplier=payload.supplier,
        airline=payload.airline,
        fare_type=payload.fare_type,
        booking_fare=fare,
        booking_date=booking_dt,
        travel_date=travel_dt,
        status=payload.status or "Completed",
        airline_pnr=payload.airline_pnr,
        booking_type=payload.booking_type,
        sector=payload.sector,
        parent_pnr=payload.parent_pnr,
        pax_name=payload.pax_name,
        source_status=payload.source_status or payload.status or "Completed",
        username=payload.username
    )
    db.add(booking)
    db.commit()
    db.refresh(booking)

    client_name = cust.name if cust else cust_name
    return schemas.BookingResponse(
        ref=booking.booking_ref,
        booking_ref=booking.booking_ref,
        customer=booking.customer_id,
        customer_id=booking.customer_id,
        customer_name=client_name,
        client=client_name,
        supplier=booking.supplier,
        airline=booking.airline,
        fare_type=booking.fare_type,
        fare=float(booking.booking_fare),
        booking_fare=float(booking.booking_fare),
        net_amount=float(booking.booking_fare),
        travel=booking.travel_date.strftime("%Y-%m-%d"),
        travel_date=booking.travel_date.isoformat(),
        booking_date=booking.booking_date.strftime("%Y-%m-%d %H:%M") if booking.booking_date else "",
        booked_date=booking.booking_date.strftime("%Y-%m-%d %H:%M") if booking.booking_date else "",
        status=booking.status,
        airline_pnr=booking.airline_pnr,
        booking_type=booking.booking_type,
        sector=booking.sector,
        parent_pnr=booking.parent_pnr,
        pax_name=booking.pax_name,
        source_status=booking.source_status,
        username=booking.username
    )


# --- Redemptions ---
@app.get("/api/redemptions", tags=["Management"])
def list_redemptions(db: Session = Depends(get_db)):
    """Get all coupon redemptions"""
    redemptions = db.query(models.CouponRedemption).order_by(desc(models.CouponRedemption.created_at)).all()
    return [
        {
            "id": r.redemption_id,
            "redemption_id": r.redemption_id,
            "customer": r.customer_id,
            "customer_id": r.customer_id,
            "booking": r.booking_ref,
            "booking_ref": r.booking_ref,
            "amount": float(r.amount_redeemed),
            "amount_redeemed": float(r.amount_redeemed),
            "booking_fare": float(r.booking_fare),
            "customer_payable": float(r.customer_payable),
            "status": r.status,
            "date": r.created_at.strftime("%Y-%m-%d") if r.created_at else ""
        }
        for r in redemptions
    ]


# --- Coupon Rules ---
@app.get("/api/rules", response_model=List[schemas.RuleResponse], tags=["Rules"])
def list_rules(db: Session = Depends(get_db)):
    """Get all coupon calculation rules ordered by priority"""
    rules = db.query(models.CouponRule).order_by(desc(models.CouponRule.priority), models.CouponRule.id.asc()).all()
    return [
        schemas.RuleResponse(
            id=r.id,
            rule_id=r.rule_id,
            office_id=r.office_id or "-",
            booking_type=r.booking_type or "-",
            supplier=r.supplier or "-",
            airline=r.airline or "-",
            fare=r.fare_type or "-",
            fare_type=r.fare_type or "-",
            percent=float(r.coupon_percent),
            coupon_percent=float(r.coupon_percent),
            priority=r.priority,
            status=r.status
        )
        for r in rules
    ]

@app.get("/api/rules/stats", tags=["Rules"])
def get_rules_stats(db: Session = Depends(get_db)):
    """
    Get detailed breakdown of rules engine, live lifecycle metrics,
    and explanations of how coupon calculations operate.
    """
    from sqlalchemy import func
    
    # Auto mature pending coupons first
    auto_mature_pending_coupons(db)

    active_rules_count = db.query(func.count(models.CouponRule.rule_id)).filter(models.CouponRule.status == "Active").scalar() or 0
    total_rules_count = db.query(func.count(models.CouponRule.rule_id)).scalar() or 0

    balance_agg = db.query(
        func.sum(models.CouponBalance.available).label("available"),
        func.sum(models.CouponBalance.pending).label("pending"),
        func.sum(models.CouponBalance.total_earned).label("total_earned"),
        func.sum(models.CouponBalance.redeemed).label("redeemed"),
    ).first()

    available = float(balance_agg.available or 0.0) if balance_agg else 0.0
    pending = float(balance_agg.pending or 0.0) if balance_agg else 0.0
    total_earned = float(balance_agg.total_earned or 0.0) if balance_agg else 0.0
    redeemed = float(balance_agg.redeemed or 0.0) if balance_agg else 0.0

    rules = db.query(models.CouponRule).order_by(desc(models.CouponRule.priority), models.CouponRule.id.asc()).all()
    
    rules_detail = [
        {
            "id": r.id,
            "rule_id": r.rule_id,
            "office_id": r.office_id or "Any Office",
            "booking_type": r.booking_type or "Any Type",
            "supplier": r.supplier or "Any Supplier",
            "airline": r.airline or "Any Airline",
            "fare_type": r.fare_type or "Any Fare",
            "coupon_percent": float(r.coupon_percent),
            "priority": r.priority,
            "status": r.status,
            "match_condition": f"{'Office: ' + r.office_id + ' | ' if r.office_id else ''}{'Type: ' + r.booking_type + ' | ' if r.booking_type else ''}{'Airline: ' + r.airline + ' | ' if r.airline else ''}{'Fare: ' + r.fare_type if r.fare_type else 'Default Tier'}"
        }
        for r in rules
    ]

    lifecycle_logic = {
        "rule_matching": "Evaluated top-down by Priority (Higher = First match wins). Checks Office ID -> Booking Type -> Supplier -> Airline -> Fare Type. If no custom match, falls back to Default Rule.",
        "pending_maturity": "Coupons earned on bookings start in 'Pending' status. They are safely locked during travel until Maturity Date (Travel Date + 1 Day).",
        "auto_credit": "Once Travel Date + 1 Day is reached, the system automatically matures and credits the coupon into the customer's 'Available' balance on their client portal without manual action.",
        "available_coupons": "Matured coupons with status 'Eligible' are available for redemption against future passenger booking fares.",
        "total_redeemed": "Coupons applied to discount new bookings. Deducted from Available balance and recorded in the immutable audit ledger."
    }

    return {
        "status": "success",
        "active_rules": active_rules_count,
        "total_rules": total_rules_count,
        "total_earned": total_earned,
        "pending_maturity": pending,
        "available_coupons": available,
        "total_redeemed": redeemed,
        "rules": rules_detail,
        "lifecycle_logic": lifecycle_logic
    }

@app.post("/api/rules", response_model=schemas.RuleResponse, tags=["Rules"])
def create_rule(payload: schemas.RuleCreate, db: Session = Depends(get_db)):
    """Create a new coupon rule with Office ID, Booking Type, Airline, and auto-priority specificity support"""
    rule_id = payload.rule_id or f"RULE-{int(time.time())}"
    percent = payload.percent if payload.percent is not None else (payload.coupon_percent or 1.0)
    
    off_id = payload.office_id if payload.office_id and payload.office_id != "-" else None
    b_type = payload.booking_type if payload.booking_type and payload.booking_type != "-" else None
    supp = payload.supplier if payload.supplier and payload.supplier != "-" else None
    air = payload.airline if payload.airline and payload.airline != "-" else None
    fare_t = payload.fare_type if payload.fare_type and payload.fare_type != "-" else None

    # Calculate specificity priority if not explicitly specified
    computed_pri = compute_rule_priority(
        office_id=off_id,
        booking_type=b_type,
        supplier=supp,
        airline=air,
        fare_type=fare_t,
        explicit_priority=payload.priority
    )
    
    rule = models.CouponRule(
        rule_id=rule_id,
        office_id=off_id,
        booking_type=b_type,
        supplier=supp,
        airline=air,
        fare_type=fare_t,
        coupon_percent=percent,
        priority=computed_pri,
        status=payload.status or "Active"
    )
    db.add(rule)
    db.commit()
    db.refresh(rule)

    return schemas.RuleResponse(
        id=rule.id,
        rule_id=rule.rule_id,
        office_id=rule.office_id or "-",
        booking_type=rule.booking_type or "-",
        supplier=rule.supplier or "-",
        airline=rule.airline or "-",
        fare=rule.fare_type or "-",
        fare_type=rule.fare_type or "-",
        percent=float(rule.coupon_percent),
        coupon_percent=float(rule.coupon_percent),
        priority=rule.priority,
        status=rule.status
    )


@app.post("/api/rules/match", tags=["Rules"])
@app.get("/api/rules/match", tags=["Rules"])
def match_rule_api(
    office_id: Optional[str] = None,
    booking_type: Optional[str] = None,
    supplier: Optional[str] = None,
    airline: Optional[str] = None,
    fare_type: Optional[str] = None,
    fare: Optional[float] = 10000.0,
    payload: Optional[dict] = Body(default=None),
    db: Session = Depends(get_db)
):
    """
    Test / simulate coupon rule matching for given booking parameters.
    Returns the matching rule, matched scope, priority, coupon % and calculated coupon.
    """
    rules = db.query(models.CouponRule).filter(models.CouponRule.status == "Active").order_by(
        desc(models.CouponRule.priority),
        models.CouponRule.id.asc()
    ).all()

    p_off = (payload.get("office_id") if payload else None) or office_id
    p_bt = (payload.get("booking_type") if payload else None) or booking_type
    p_supp = (payload.get("supplier") if payload else None) or supplier
    p_air = (payload.get("airline") if payload else None) or airline
    p_fare_t = (payload.get("fare_type") if payload else None) or fare_type
    p_fare = float((payload.get("booking_fare") or payload.get("fare") if payload else None) or fare or 10000.0)

    matched_rule = find_matching_rule(
        rules=rules,
        office_id=p_off,
        booking_type=p_bt,
        supplier=p_supp,
        airline=p_air,
        fare_type=p_fare_t
    )

    if matched_rule:
        percent = float(matched_rule.coupon_percent)
        rule_id = matched_rule.rule_id
        priority = matched_rule.priority
        match_scope = f"{'Airline: ' + matched_rule.airline if matched_rule.airline else 'Any Airline'}{' | Fare: ' + matched_rule.fare_type if matched_rule.fare_type else ' | Any Fare'}"
        is_default = (not matched_rule.airline and not matched_rule.fare_type and not matched_rule.office_id and not matched_rule.booking_type and not matched_rule.supplier)
    else:
        percent = 1.0
        rule_id = "DEFAULT-FALLBACK"
        priority = 0
        match_scope = "Default Tier Fallback"
        is_default = True

    coupon_amount = round(p_fare * (percent / 100.0), 2)
    customer_payable = round(max(0.0, p_fare - coupon_amount), 2)

    return {
        "status": "success",
        "matched_rule_id": rule_id,
        "priority": priority,
        "coupon_percent": percent,
        "coupon_amount": coupon_amount,
        "booking_fare": p_fare,
        "customer_payable": customer_payable,
        "match_scope": match_scope,
        "is_default_fallback": is_default,
        "query": {
            "office_id": p_off,
            "booking_type": p_bt,
            "supplier": p_supp,
            "airline": p_air,
            "fare_type": p_fare_t
        }
    }



# --- System Settings ---
@app.get("/api/settings", response_model=schemas.SettingsSchema, tags=["Settings"])
def get_settings(db: Session = Depends(get_db)):
    """Get global configuration parameters"""
    settings_records = db.query(models.SystemSetting).all()
    settings_map = {s.config_key: s.config_value for s in settings_records}

    return schemas.SettingsSchema(
        min_redemption=float(settings_map.get("min_redemption", 100)),
        max_redemption=float(settings_map.get("max_redemption", 50000)),
        expiry_days=int(settings_map.get("expiry_days", 365)),
        allow_partial_redemption=settings_map.get("allow_partial_redemption", "true").lower() == "true",
        allow_combined_offers=settings_map.get("allow_combined_offers", "false").lower() == "true"
    )

@app.post("/api/settings", response_model=schemas.SettingsSchema, tags=["Settings"])
def update_settings(payload: schemas.SettingsSchema, db: Session = Depends(get_db)):
    """Update global configuration parameters in MySQL"""
    items = {
        "min_redemption": str(payload.min_redemption),
        "max_redemption": str(payload.max_redemption),
        "expiry_days": str(payload.expiry_days),
        "allow_partial_redemption": "true" if payload.allow_partial_redemption else "false",
        "allow_combined_offers": "true" if payload.allow_combined_offers else "false"
    }

    for k, v in items.items():
        rec = db.query(models.SystemSetting).filter(models.SystemSetting.config_key == k).first()
        if rec:
            rec.config_value = v
        else:
            db.add(models.SystemSetting(config_key=k, config_value=v))

    db.commit()
    return payload



# ============================================================================
# EXCEL UPLOAD & TEMPLATE — BULK BOOKING & PASSENGER IMPORT
# ============================================================================

EXCEL_COLUMNS = [
    "booking_date", "customer_name", "customer_id", "booking_ref",
    "airline_pnr", "source_status", "booking_type", "username",
    "pax_name", "sector", "travel_date", "parent_pnr", "booking_fare",
    "supplier", "airline", "fare_type"
]

# Maps each canonical internal field to the list of possible Excel header names
# (lowercased & whitespace normalized). This recognizes all variations including:
# Booked Date | Client | Client ID | S PNR | Airline PNR | Status | Booking Type | Username | Pax Name | Sector | Date of Travel | Parent PNR | Net Amount
# as well as PNRWiseSales: First Name | Last Name | Pax Type | Net Amount
HEADER_ALIASES = {
    "booking_date":  ["booked date", "booking_date", "booking date", "booked_date", "booked on", "date", "booking date & time", "booked date & time", "booked dt", "txn date", "issue date"],
    "customer_name": ["client", "customer_name", "customer name", "client name", "client_name", "agency", "agency name", "company", "company name", "party name", "account name"],
    "customer_id":   ["client id", "client_id", "customer_id", "customer id", "client code", "clientcode", "cust id", "cust_id", "clientid", "account id", "agent code", "party code"],
    "office_id":     ["office id", "office_id", "officeid", "office", "branch", "branch id", "branch code", "office code", "location"],
    "booking_ref":   ["s pnr", "s_pnr", "spnr", "booking_ref", "booking ref", "pnr", "system pnr", "system_pnr", "pnr no", "pnr number", "reference"],
    "airline_pnr":   ["airline pnr", "airline_pnr", "airlinepnr", "air pnr", "air_pnr", "carrier pnr", "gds pnr", "airline ref"],
    "source_status": ["status", "source_status", "source status", "booking status", "pnr status"],
    "booking_type":  ["booking type", "booking_type", "trip type", "trip_type", "type", "journey type", "bookingtype", "btype"],
    "username":      ["username", "user name", "user", "agent", "booked by", "agent name", "staff", "consultant"],
    "pax_name":      ["pax name", "pax_name", "pax", "passenger", "passenger name", "passenger_name", "name", "pax details"],
    "first_name":    ["first name", "firstname", "first_name", "fname", "given name", "pax first name"],
    "last_name":     ["last name", "lastname", "last_name", "lname", "surname", "pax last name"],
    "pax_type":      ["pax type", "pax_type", "paxtype", "passenger type", "type of pax"],
    "sector":        ["sector", "sectors", "route", "routing", "origin destination", "orig-dest", "itinerary"],
    "travel_date":   ["date of travel", "date_of_travel", "travel_date", "travel date", "traveldate", "departure date", "dep date", "travel dt", "travel_dt", "journey date", "travel"],
    "parent_pnr":    ["parent pnr", "parent_pnr", "parentpnr", "main pnr", "group pnr", "original pnr"],
    "booking_fare":  ["net amount", "net_amount", "netamount", "booking_fare", "booking fare", "fare", "amount", "total amount", "net fare", "net_fare", "gross amount", "ticket cost", "total fare", "total"],
    "supplier":      ["supplier", "vendor", "provider"],
    "airline":       ["airline", "carrier", "airline name"],
    "fare_type":     ["fare_type", "fare type", "fare category", "faretype"],
}

def parse_flexible_date(date_val, default=None):
    """
    Parses a wide variety of date inputs from Excel or JSON:
    - datetime / date object
    - ISO format 'YYYY-MM-DD' or 'YYYY-MM-DDTHH:MM:SS'
    - 'DD/MM/YYYY' or 'DD/MM/YYYY HH:MM' or 'DD/MM/YYYY HH:MM:SS'
    - 'DD-MM-YYYY' or 'YYYY/MM/DD'
    - Multi-leg travel date '09/10/2026-10/10/2026-04/11/2026' (uses first leg)
    - Excel serial number float/int (e.g. 45533)
    """
    if date_val is None:
        return default
    if isinstance(date_val, datetime):
        return date_val.replace(tzinfo=None)
    if hasattr(date_val, "year") and hasattr(date_val, "month") and hasattr(date_val, "day"):
        return datetime(date_val.year, date_val.month, date_val.day)
    if isinstance(date_val, (int, float)):
        try:
            # Excel base date is 1899-12-30
            return datetime(1899, 12, 30) + timedelta(days=float(date_val))
        except Exception:
            return default

    s = str(date_val).strip()
    if not s:
        return default

    # Handle multi-leg string like '09/10/2026-10/10/2026-04/11/2026' or '31/08/2026-31/08/2026'
    if "-" in s and "/" in s and len(s) > 10:
        s = s.split("-")[0].strip()

    s_clean = s.replace("Z", "").strip()

    # Try ISO first
    try:
        return datetime.fromisoformat(s_clean)
    except ValueError:
        pass

    # Try common travel date/datetime formats
    formats = [
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y",
        "%m-%d-%Y",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s_clean, fmt)
        except ValueError:
            pass

    return default

def _process_single_booking_row(row_data: dict, db: Session) -> dict:
    """
    Core logic: process one booking row from the Excel file.
    Intelligently extracts Client ID (e.g. RCMAA0500655), Client Name,
    combines First Name + Last Name, generates fallback PNR if missing,
    and earns coupons for the client in MySQL.
    """
    row_num = row_data.get("_row", "?")
    try:
        customer_id_raw  = str(row_data.get("customer_id", "") or "").strip()
        customer_name_raw= str(row_data.get("customer_name", "") or "").strip()
        first_name       = str(row_data.get("first_name", "") or "").strip()
        last_name        = str(row_data.get("last_name", "") or "").strip()
        pax_name_raw     = str(row_data.get("pax_name", "") or "").strip()
        booking_ref_raw  = str(row_data.get("booking_ref", "") or "").strip()
        supplier         = str(row_data.get("supplier", "") or "").strip() or None
        airline          = str(row_data.get("airline", "") or "").strip() or None
        fare_type        = str(row_data.get("fare_type", "") or "").strip() or None
        booking_fare_raw = row_data.get("booking_fare", 0)
        travel_date_raw  = row_data.get("travel_date", "")
        booking_date_raw = row_data.get("booking_date")
        airline_pnr      = str(row_data.get("airline_pnr", "") or "").strip() or None
        office_id        = str(row_data.get("office_id", "") or "").strip() or None
        booking_type     = str(row_data.get("booking_type", "") or "").strip() or None
        sector           = str(row_data.get("sector", "") or "").strip() or None
        parent_pnr       = str(row_data.get("parent_pnr", "") or "").strip() or None
        source_status    = str(row_data.get("source_status", "") or "").strip() or "Confirmed"
        username         = str(row_data.get("username", "") or "").strip() or None

        # ---- 1. Smart Client ID & Client Name Extraction ----
        customer_id = customer_id_raw
        customer_name = customer_name_raw

        # If Client ID is in brackets in the Client field (e.g. "SAKSOFT LIMITED [RCMAA0500655]")
        if not customer_id and customer_name:
            bracket_match = re.search(r'\[([A-Za-z0-9_-]+)\]', customer_name)
            if bracket_match:
                customer_id = bracket_match.group(1).strip()

        # Clean customer_name to remove bracketed client IDs
        if customer_name:
            customer_name = re.sub(r'\s*\[.*?\]\s*', '', customer_name).strip()

        # If customer_id is still missing, derive clean code from customer_name
        if not customer_id and customer_name:
            derived = re.sub(r'[^A-Za-z0-9]', '', customer_name)[:16].upper()
            customer_id = f"CL-{derived}" if derived else f"CL-{row_num}"

        if not customer_id:
            customer_id = f"CL-{row_num}"

        if not customer_name:
            customer_name = f"Client {customer_id}"

        # ---- 2. Smart Pax Name Assembly ----
        pax_name = pax_name_raw
        if not pax_name and (first_name or last_name):
            pax_name = f"{first_name} {last_name}".strip()
        if not pax_name:
            pax_name = customer_name or f"Pax-{row_num}"

        # ---- 3. Smart Booking Reference (S PNR) ----
        booking_ref = booking_ref_raw
        if not booking_ref:
            if airline_pnr:
                booking_ref = airline_pnr
            else:
                booking_ref = f"PNR-{customer_id}-{row_num}"

        # ---- 4. Clean & Parse booking_fare (Net Amount) ----
        try:
            if isinstance(booking_fare_raw, str):
                cleaned_fare = re.sub(r'[^\d.]', '', booking_fare_raw.replace(",", ""))
                booking_fare = float(cleaned_fare) if cleaned_fare else 0.0
            else:
                booking_fare = float(booking_fare_raw or 0)
            if booking_fare <= 0:
                raise ValueError("Net Amount must be greater than 0")
        except Exception:
            return {
                "row": row_num,
                "status": "failed",
                "booking_ref": booking_ref,
                "pax_name": pax_name or "-",
                "customer_id": customer_id,
                "message": f"Invalid Net Amount: '{booking_fare_raw}' (must be > 0)"
            }

        # ---- 5. Parse travel_date & booking_date ----
        travel_dt = parse_flexible_date(travel_date_raw)
        if not travel_dt:
            # Fallback to booking_date or today + 7 days
            parsed_booking_date_tmp = parse_flexible_date(booking_date_raw)
            travel_dt = parsed_booking_date_tmp if parsed_booking_date_tmp else datetime.utcnow()

        parsed_booking_date = parse_flexible_date(booking_date_raw, default=datetime.utcnow())

        # ---- 6. Check duplicate booking_ref ----
        existing = db.query(models.Booking).filter(
            models.Booking.booking_ref == booking_ref
        ).first()
        if existing:
            return {
                "row": row_num,
                "status": "skipped",
                "booking_ref": booking_ref,
                "customer_id": customer_id,
                "pax_name": pax_name or existing.pax_name or "-",
                "client": customer_name,
                "customer_name": customer_name,
                "s_pnr": booking_ref,
                "net_amount": booking_fare,
                "message": f"Booking Ref (S PNR) '{booking_ref}' already exists (skipped duplicate)"
            }

        # ---- 7. Ensure Customer & CouponBalance exist in MySQL ----
        customer = db.query(models.Customer).filter(
            models.Customer.customer_id == customer_id
        ).first()
        if not customer:
            email_val = f"{customer_id.lower()}@travel.com"
            if username and "@" in username:
                email_val = username
            elif "@" in customer_name_raw:
                email_val = customer_name_raw.strip()

            customer = models.Customer(
                customer_id=customer_id,
                name=customer_name or f"Client {customer_id}",
                email=email_val,
                status="Active"
            )
            db.add(customer)
            db.commit()
        elif customer_name and customer.name.startswith("Client ") and customer_name != f"Client {customer_id}":
            customer.name = customer_name
            db.commit()

        # ---- 8. Priority Rule Matching ----
        rules = db.query(models.CouponRule).filter(
            models.CouponRule.status == "Active"
        ).order_by(desc(models.CouponRule.priority), models.CouponRule.id.asc()).all()

        matched_rule = find_matching_rule(
            rules=rules,
            office_id=office_id,
            booking_type=booking_type,
            supplier=supplier,
            airline=airline,
            fare_type=fare_type
        )
        matched_percent = float(matched_rule.coupon_percent) if matched_rule else 1.0

        coupon_earned = round(booking_fare * (matched_percent / 100.0), 2)
        eligibility_dt = travel_dt + timedelta(days=1)

        # Normalize Ticketed status to Confirmed
        norm_status = (source_status or "Confirmed").strip()
        if norm_status.lower() in ["ticketed", "ticket"]:
            norm_status = "Confirmed"

        # ---- 9. Save Booking ----
        booking = models.Booking(
            booking_ref=booking_ref,
            customer_id=customer_id,
            supplier=supplier,
            airline=airline,
            fare_type=fare_type,
            booking_fare=booking_fare,
            booking_date=parsed_booking_date,
            travel_date=travel_dt,
            status=norm_status,
            airline_pnr=airline_pnr,
            booking_type=booking_type,
            sector=sector,
            parent_pnr=parent_pnr,
            pax_name=pax_name,
            source_status=norm_status,
            username=username
        )
        db.add(booking)

        # ---- 10. Save Coupon (Pending until travel date) ----
        coupon_id = f"CPN-{booking_ref}"
        coupon = models.Coupon(
            coupon_id=coupon_id,
            booking_ref=booking_ref,
            customer_id=customer_id,
            coupon_amount=coupon_earned,
            coupon_percent=matched_percent,
            status="Pending",
            eligibility_date=eligibility_dt,
            created_at=datetime.utcnow()
        )
        db.add(coupon)

        # ---- 11. Write to Ledger ----
        txn_id = f"TXN-UPLOAD-{int(time.time())}.{random.randint(1,999)}"
        ledger_entry = models.CouponLedger(
            txn_id=txn_id,
            customer_id=customer_id,
            booking_ref=booking_ref,
            txn_type="Coupon Earned",
            booking_fare=booking_fare,
            coupon_percent=matched_percent,
            coupon_earned=coupon_earned,
            amount=coupon_earned,
            status="Pending",
            travel_date=travel_dt,
            created_at=datetime.utcnow()
        )
        db.add(ledger_entry)

        # ---- 12. Update Coupon Balance in MySQL ----
        balance = db.query(models.CouponBalance).filter(
            models.CouponBalance.customer_id == customer_id
        ).first()

        if not balance:
            balance = models.CouponBalance(
                customer_id=customer_id,
                total_earned=coupon_earned,
                pending=coupon_earned,
                available=0.0,
                redeemed=0.0,
                expired=0.0,
                cancelled=0.0
            )
            db.add(balance)
        else:
            balance.total_earned = float(balance.total_earned or 0.0) + coupon_earned
            balance.pending = float(balance.pending or 0.0) + coupon_earned

        db.commit()

        return {
            "row": row_num,
            "status": "success",
            "booking_ref": booking_ref,
            "customer_id": customer_id,
            "pax_name": pax_name,
            "client": customer_name,
            "customer_name": customer_name,
            "s_pnr": booking_ref,
            "sector": sector or "-",
            "net_amount": booking_fare,
            "booking_fare": booking_fare,
            "coupon_percent": matched_percent,
            "coupon_earned": coupon_earned,
            "eligibility_date": eligibility_dt.strftime("%Y-%m-%d"),
            "message": f"Successfully processed. Coupon ₹{coupon_earned} ({matched_percent}%) earned for {customer_name}"
        }

    except Exception as e:
        db.rollback()
        return {
            "row": row_num,
            "status": "failed",
            "booking_ref": row_data.get("booking_ref", "N/A"),
            "customer_id": row_data.get("customer_id", "N/A"),
            "message": f"Database error: {str(e)}"
        }


# ============================================================================
# EXCEL UPLOAD API ENDPOINTS
# ============================================================================
@app.post("/api/upload/excel", tags=["Excel Upload"])
async def upload_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Bulk upload bookings and passengers from an Excel (.xlsx) file.
    Validates headers, extracts passenger & client details,
    calculates coupons via rule hierarchy, and updates MySQL DB.
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file format. Please upload an Excel (.xlsx) file."
        )

    try:
        contents = await file.read()
        import io
        wb = openpyxl.load_workbook(filename=io.BytesIO(contents), data_only=True)
        ws = wb.active

        # Extract headers from row 1
        headers = []
        for cell in ws[1]:
            val = str(cell.value or "").strip()
            headers.append(val)

        # Map Excel columns to internal keys
        header_mapping = {
            "booked date": "booking_date",
            "booked_date": "booking_date",
            "booking date": "booking_date",
            "client": "customer_name",
            "client name": "customer_name",
            "customer name": "customer_name",
            "client id": "customer_id",
            "client_id": "customer_id",
            "customer id": "customer_id",
            "customer_id": "customer_id",
            "office id": "office_id",
            "office_id": "office_id",
            "office": "office_id",
            "branch": "office_id",
            "branch id": "office_id",
            "s pnr": "booking_ref",
            "s_pnr": "booking_ref",
            "spnr": "booking_ref",
            "booking ref": "booking_ref",
            "booking_ref": "booking_ref",
            "airline pnr": "airline_pnr",
            "airline_pnr": "airline_pnr",
            "pnr": "airline_pnr",
            "status": "source_status",
            "booking type": "booking_type",
            "booking_type": "booking_type",
            "bookingtype": "booking_type",
            "btype": "booking_type",
            "trip type": "booking_type",
            "username": "username",
            "agent": "username",
            "pax name": "pax_name",
            "pax_name": "pax_name",
            "passenger name": "pax_name",
            "passenger": "pax_name",
            "first name": "first_name",
            "last name": "last_name",
            "sector": "sector",
            "route": "sector",
            "date of travel": "travel_date",
            "travel date": "travel_date",
            "travel_date": "travel_date",
            "parent pnr": "parent_pnr",
            "parent_pnr": "parent_pnr",
            "net amount": "booking_fare",
            "net_amount": "booking_fare",
            "fare": "booking_fare",
            "amount": "booking_fare",
            "supplier": "supplier",
            "airline": "airline",
            "fare type": "fare_type",
            "fare_type": "fare_type"
        }

        col_key_map = {}
        for col_idx, h in enumerate(headers):
            h_norm = h.lower().replace("_", " ").strip()
            if h_norm in header_mapping:
                col_key_map[col_idx] = header_mapping[h_norm]

        results = []
        success_count = 0
        failed_count = 0
        skipped_count = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            row_dict = {"_row": row_idx}
            for col_idx, val in enumerate(row):
                if col_idx in col_key_map:
                    row_dict[col_key_map[col_idx]] = val

            res = _process_single_booking_row(row_dict, db)
            results.append(res)
            if res["status"] == "success":
                success_count += 1
            elif res["status"] == "skipped":
                skipped_count += 1
            else:
                failed_count += 1

        return {
            "status": "success",
            "summary": {
                "total_rows": len(results),
                "success": success_count,
                "skipped": skipped_count,
                "failed": failed_count
            },
            "results": results
        }

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error parsing Excel file: {str(e)}"
        )


@app.get("/api/upload/template", tags=["Excel Upload"])
def download_template():
    """
    Download a ready-to-use Excel template for bulk booking & passenger upload.
    Includes the 13 standard columns and sample confirmed bookings.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bookings"

    # Header row (13 columns)
    headers = [
        "Booked Date", "Client", "Client ID", "S PNR", "Airline PNR",
        "Status", "Booking Type", "Username", "Pax Name", "Sector",
        "Date of Travel", "Parent PNR", "Net Amount"
    ]
    ws.append(headers)

    # Styling
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left", vertical="center")
    right_align  = Alignment(horizontal="right", vertical="center")
    thin_border  = Border(
        left=Side(style="thin", color="CBD5E1"), right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"), bottom=Side(style="thin", color="CBD5E1")
    )

    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border
    ws.row_dimensions[1].height = 28

    # Column widths
    col_widths = [18, 20, 14, 16, 16, 14, 16, 16, 24, 18, 16, 16, 14]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Sample realistic data rows with Confirmed status and diverse booking types
    sample_rows = [
        ["2026-08-29 10:30", "ABC Holidays",   "CL-1001", "SPNR78901", "6E-XY789", "Confirmed", "Web Booking",           "agent_priya",  "MR RAHUL SHARMA",     "DEL-BOM-DEL", "2026-09-15", "",          12500],
        ["2026-08-29 11:15", "Global Travels", "CL-1002", "SPNR78902", "AI-AB456", "Confirmed", "Indesk Booking",        "agent_vikram", "MS ANITHA RAJAN",     "MAA-BLR",     "2026-09-20", "",           6800],
        ["2026-08-29 11:45", "ABC Holidays",   "CL-1001", "SPNR78903", "SG-98765", "Confirmed", "Travel Desk Booking",   "agent_priya",  "MR KARTHIK SUNDAR",   "BLR-DEL-BLR", "2026-10-05", "SPNR78901", 18400],
        ["2026-08-29 12:00", "Star Tours",     "CL-1003", "SPNR78904", "UK-54321", "Confirmed", "Mobile Booking",        "agent_deepa",  "MRS POOJA HEGDE",     "BOM-GOI-BOM", "2026-10-12", "",          14200],
        ["2026-08-29 12:30", "Global Travels", "CL-1002", "SPNR78905", "6E-55443", "Confirmed", "Retrieve PNR Ticketing","agent_vikram", "MR SURESH MENON",     "MAA-HYD",     "2026-11-01", "",           5400],
    ]

    fill_even = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_odd  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for r_idx, sample in enumerate(sample_rows, start=2):
        ws.append(sample)
        ws.row_dimensions[r_idx].height = 22
        for col_i, cell in enumerate(ws[r_idx], start=1):
            cell.border = thin_border
            cell.fill   = fill_even if r_idx % 2 == 0 else fill_odd
            if col_i in (2, 9):  # Client & Pax Name
                cell.alignment = left_align
            elif col_i == 13:   # Net Amount
                cell.alignment = right_align
                cell.number_format = '#,##0.00'
            else:
                cell.alignment = center_align

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    ws2["A1"] = "HOW TO USE THIS PASSENGER & BOOKINGS TEMPLATE"
    ws2["A1"].font = Font(bold=True, size=14, color="1E3A8A")
    instructions = [
        ("", ""),
        ("COLUMN", "DESCRIPTION & RULES", "REQUIRED / OPTIONAL"),
        ("Booked Date", "Date & time the booking was made (e.g. 2026-08-29 10:30 or 29/08/2026). Defaults to now if empty.", "Optional"),
        ("Client", "Name of the travel agency or corporate client (e.g. ABC Holidays).", "Optional (Auto-created)"),
        ("Client ID", "Unique Client/Customer ID (e.g. CL-1001 or CUST001). Must exist or will be auto-created.", "REQUIRED"),
        ("S PNR", "System Booking Reference / PNR (e.g. SPNR78901). Duplicate PNRs in DB will be skipped.", "REQUIRED"),
        ("Airline PNR", "Airline confirmation PNR code (e.g. 6E-XY789, AI-AB456).", "Optional"),
        ("Status", "Booking / Ticket status (e.g. Ticketed, Confirmed, Completed).", "Optional"),
        ("Booking Type", "Trip type: Round Trip, One Way, Multi City, etc.", "Optional"),
        ("Username", "Booking agent / staff username who created the booking.", "Optional"),
        ("Pax Name", "Passenger full name (e.g. MR RAHUL SHARMA).", "Optional / Recommended"),
        ("Sector", "Flight sector route (e.g. DEL-BOM-DEL or MAA-BLR).", "Optional / Recommended"),
        ("Date of Travel", "Travel date (e.g. 2026-09-15, 15/09/2026, or multi-leg 09/10/2026-10/10/2026).", "REQUIRED"),
        ("Parent PNR", "Original parent PNR if this booking was split or linked.", "Optional"),
        ("Net Amount", "Total net ticket fare amount in INR (e.g. 12500). Must be a positive number.", "REQUIRED"),
        ("", "", ""),
        ("COUPON LIFECYCLE NOTES:", "", ""),
        ("", "- Coupons are automatically calculated based on your Active Coupon Rules (default 1%).", ""),
        ("", "- All created coupons start in Pending status.", ""),
        ("", "- Coupons become Eligible for redemption 1 day after the Date of Travel.", ""),
        ("", "- All entries are recorded in the Real-Time Transaction Ledger and Customer Balances.", ""),
    ]
    for row_data in instructions:
        ws2.append(list(row_data))
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 80
    ws2.column_dimensions["C"].width = 25

    # Style header in instructions
    ws2["A3"].font = Font(bold=True, color="1E3A8A")
    ws2["B3"].font = Font(bold=True, color="1E3A8A")
    ws2["C3"].font = Font(bold=True, color="1E3A8A")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    headers_response = {
        "Content-Disposition": 'attachment; filename="Roundtrip_COUPON_template.xlsx"',
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }
    return StreamingResponse(buf, media_type=headers_response["Content-Type"], headers=headers_response)


# ============================================================================
# ROOT ROUTE
# ============================================================================
@app.get("/", tags=["General"])
def root():
    return {
        "message": "Coupon Management System Backend API is running with MySQL database.",
        "docs_url": "/docs",
        "health_check": "/api/health"
    }


# ============================================================================
# MAIN ENTRYPOINT
# ============================================================================
if __name__ == "__main__":
    port = int(os.getenv("PORT", "8000"))
    host = os.getenv("HOST", "0.0.0.0")
    print(f"[INFO] Starting Coupon Management System on http://127.0.0.1:{port}")
    uvicorn.run("main:app", host=host, port=port, reload=True)
