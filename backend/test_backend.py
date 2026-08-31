import os
import sys
import asyncio

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

import httpx
from main import app

async def run_all_tests():
    print("==================================================")
    print("  RUNNING COUPON BACKEND TEST SUITE WITH MYSQL")
    print("==================================================")
    
    transport = httpx.ASGITransport(app=app)
    async with httpx.AsyncClient(transport=transport, base_url="http://test") as client:
        # 1. Health Check
        r = await client.get("/api/health")
        assert r.status_code == 200, f"Health check failed: {r.text}"
        data = r.json()
        assert data["status"] == "healthy"
        assert data["database"] == "connected"
        print(f"[PASS] 1. Health check passed: {data}")

        # 2. Get Balance
        r = await client.get("/api/coupon/balance/CUST001")
        assert r.status_code == 200, f"Balance failed: {r.text}"
        data = r.json()
        assert data["customer_id"] == "CUST001"
        assert "available" in data
        assert "pending" in data
        print(f"[PASS] 2. Customer Balance: Available=₹{data['available']}, Pending=₹{data['pending']}, Earned=₹{data['total_earned']}")

        # 3. Earn Coupon
        earn_payload = {
            "customer_id": "CUST001",
            "customer_name": "Rajesh Kumar",
            "booking_ref": "BK-TEST-ASYNC-01",
            "supplier": "Supplier A",
            "airline": "IndiGo",
            "fare_type": "Super 6E",
            "booking_fare": 10000.0,
            "travel_date": "2026-12-10T14:30:00Z",
            "booking_date": "2026-08-27T10:00:00Z"
        }
        r = await client.post("/api/coupon/earn", json=earn_payload)
        assert r.status_code == 200, f"Earn coupon failed: {r.text}"
        data = r.json()
        assert data["status"] == "success"
        assert data["coupon_percent"] == 3.0
        assert data["coupon_earned"] == 300.0
        assert data["coupon_status"] == "Pending"
        print(f"[PASS] 3. Earn Coupon: TxnID={data['txn_id']}, Earned=₹{data['coupon_earned']} ({data['coupon_percent']}%), Status={data['coupon_status']}")

        # 4. Release Coupon
        r = await client.post("/api/coupon/release", json={"booking_ref": "BK-TEST-ASYNC-01"})
        assert r.status_code == 200, f"Release coupon failed: {r.text}"
        data = r.json()
        assert data["status"] == "success"
        assert data["new_status"] == "Eligible"
        assert data["coupon_amount"] == 300.0
        print(f"[PASS] 4. Release Coupon: TxnID={data['txn_id']}, Amount=₹{data['coupon_amount']}, Status={data['new_status']}")

        # 5. Redeem Coupon
        redeem_payload = {
            "customer_id": "CUST001",
            "booking_ref": f"BK-TEST-{int(asyncio.get_event_loop().time())}",
            "amount_to_redeem": 100.0,
            "booking_fare": 5000.0
        }
        r = await client.post("/api/coupon/redeem", json=redeem_payload)
        assert r.status_code == 200, f"Redeem coupon failed: {r.text}"
        data = r.json()
        assert data["status"] == "success"
        assert data["coupon_redeemed"] == 100.0
        assert data["customer_payable"] == 4900.0
        print(f"[PASS] 5. Redeem Coupon: Redeemed=₹{data['coupon_redeemed']}, Customer Pays=₹{data['customer_payable']}, Remaining=₹{data['remaining_coupon_balance']}")

        # 6. Reverse Coupon
        await client.post("/api/coupon/earn", json={
            "customer_id": "CUST001",
            "customer_name": "Rajesh Kumar",
            "booking_ref": "BK-TEST-REVERSE-ASYNC",
            "supplier": "Supplier A",
            "airline": "IndiGo",
            "fare_type": "Super 6E",
            "booking_fare": 6000.0,
            "travel_date": "2026-12-15T14:30:00Z"
        })
        reverse_payload = {
            "original_booking_ref": "BK-TEST-REVERSE-ASYNC",
            "reason": "Cancelled",
            "remarks": "User requested flight cancellation"
        }
        r = await client.post("/api/coupon/reverse", json=reverse_payload)
        assert r.status_code == 200, f"Reverse coupon failed: {r.text}"
        data = r.json()
        assert data["status"] == "success"
        assert data["action"] == "Reversed"
        print(f"[PASS] 6. Reverse Coupon: TxnID={data['txn_id']}, Action={data['action']}, Amount=₹{data['coupon_amount']}")

        # 7. Get Ledger
        r = await client.get("/api/coupon/ledger/CUST001")
        assert r.status_code == 200, f"Get ledger failed: {r.text}"
        data = r.json()
        assert data["customer_id"] == "CUST001"
        assert len(data["ledger"]) > 0
        print(f"[PASS] 7. Ledger: Found {len(data['ledger'])} transactions for CUST001. Latest: {data['ledger'][0]['type']} - ₹{data['ledger'][0]['amount']}")

        # 8. Management CRUD Endpoints
        custs = (await client.get("/api/customers")).json()
        assert len(custs) >= 3
        print(f"[PASS] 8.1 Customers API: {len(custs)} customers loaded")

        bks = (await client.get("/api/bookings")).json()
        assert len(bks) >= 3
        print(f"[PASS] 8.2 Bookings API: {len(bks)} bookings loaded")

        rules = (await client.get("/api/rules")).json()
        assert len(rules) >= 5
        print(f"[PASS] 8.3 Rules API: {len(rules)} rules loaded")

        settings = (await client.get("/api/settings")).json()
        assert "min_redemption" in settings
        print(f"[PASS] 8.4 Settings API: min_redemption=₹{settings['min_redemption']}, max_redemption=₹{settings['max_redemption']}")

        # 9. Test 13-Column Excel Template Download
        r = await client.get("/api/upload/template")
        assert r.status_code == 200, f"Template download failed: {r.status_code}"
        import openpyxl, io
        wb = openpyxl.load_workbook(filename=io.BytesIO(r.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        expected_headers = [
            "Booked Date", "Client", "Client ID", "S PNR", "Airline PNR",
            "Status", "Booking Type", "Username", "Pax Name", "Sector",
            "Date of Travel", "Parent PNR", "Net Amount"
        ]
        assert headers == expected_headers, f"Headers mismatch: {headers}"
        print(f"[PASS] 9. Template Download: Successfully verified 13 columns -> {headers}")

        # 10. Test 13-Column Excel Upload
        upload_wb = openpyxl.Workbook()
        upload_ws = upload_wb.active
        upload_ws.append(expected_headers)
        test_pnr = f"SPNR-TEST-{int(time.time())}"
        upload_ws.append([
            "2026-08-29 10:30", "Test Agency", "CL-TEST-99", test_pnr, "6E-TEST",
            "Ticketed", "Round Trip", "agent_auto", "MR AUTO TESTER", "MAA-DEL-MAA",
            "2026-10-15", "", 15000.0
        ])
        buf = io.BytesIO()
        upload_wb.save(buf)
        buf.seek(0)

        files = {"file": ("test_upload.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = await client.post("/api/upload/excel", files=files)
        assert r.status_code == 200, f"Upload failed: {r.text}"
        res_data = r.json()
        assert res_data["summary"]["success"] == 1
        assert res_data["results"][0]["pax_name"] == "MR AUTO TESTER"
        assert res_data["results"][0]["sector"] == "MAA-DEL-MAA"
        assert res_data["results"][0]["booking_fare"] == 15000.0
        print(f"[PASS] 10. Excel 13-Column Upload: {res_data['summary']['success']} row processed, Pax='{res_data['results'][0]['pax_name']}', Coupon=₹{res_data['results'][0]['coupon_earned']}")

        # 11. Test Rules with Office ID & Booking Type and Stats
        r = await client.get("/api/rules/stats")
        assert r.status_code == 200, f"Rules stats failed: {r.text}"
        stats_data = r.json()
        assert stats_data["status"] == "success"
        assert "active_rules" in stats_data
        assert "lifecycle_logic" in stats_data
        print(f"[PASS] 11. Rules Overview & Stats: Active={stats_data['active_rules']}, Lifetime Earned=₹{stats_data['total_earned']}, Available=₹{stats_data['available_coupons']}")

        # 12. Test Auto-Maturity Endpoint
        r = await client.post("/api/coupon/auto-mature")
        assert r.status_code == 200, f"Auto mature failed: {r.text}"
        m_data = r.json()
        assert m_data["status"] == "success"
        print(f"[PASS] 12. Auto-Maturity Trigger: Matured {m_data['matured_count']} coupons.")

        # 13. Test Airline Matching Hierarchy & Any Airline Fallback
        test_cases = [
            {"airline": "IndiGo", "fare_type": "Super 6E", "expected_pct": 3.0, "name": "IndiGo Super 6E (Rule 1)"},
            {"airline": "IndiGo", "fare_type": "Normal Fare", "expected_pct": 2.0, "name": "IndiGo Any Fare (Rule 2)"},
            {"airline": "Air India", "fare_type": "Flexi", "expected_pct": 2.5, "name": "Air India Flexi (Rule 3)"},
            {"airline": "SpiceJet", "fare_type": "Regular", "expected_pct": 1.5, "name": "SpiceJet Any Fare (Rule 4)"},
            {"airline": "Akasa Air", "fare_type": "Normal Fare", "expected_pct": 1.0, "name": "Akasa Air -> Any Airline Fallback (Rule 5)"},
            {"airline": "Any Airline", "fare_type": "", "expected_pct": 1.0, "name": "Generic Any Airline (Rule 5)"},
            {"airline": "", "fare_type": "", "expected_pct": 1.0, "name": "Empty Airline -> Any Airline Fallback (Rule 5)"},
        ]

        for tc in test_cases:
            res = await client.post("/api/rules/match", json={"airline": tc["airline"], "fare_type": tc["fare_type"], "fare": 10000.0})
            assert res.status_code == 200, f"Match failed for {tc['name']}: {res.text}"
            data = res.json()
            assert data["coupon_percent"] == tc["expected_pct"], f"Expected {tc['expected_pct']}%, got {data['coupon_percent']}% for {tc['name']}"
            print(f"[PASS] 13. Match Test: {tc['name']} -> {data['coupon_percent']}% applied (Rule: {data['matched_rule_id']}, Pri: #{data['priority']})")

    print("\n==================================================")
    print("  ALL 13 TEST SUITES PASSED PERFECTLY ON MYSQL!")
    print("==================================================")

if __name__ == "__main__":
    import time
    asyncio.run(run_all_tests())

